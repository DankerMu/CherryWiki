from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base_parser import BaseParser, ParseResult
from .utils import compact_blank_lines, markdown_table, package_version


class XlsxParser(BaseParser):
    source_type = "xlsx"
    extraction_tool = "openpyxl"
    extraction_version = package_version("openpyxl")

    def parse(self, file_path: Path) -> ParseResult:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sections: list[str] = []

        try:
            for sheet in workbook.worksheets:
                sections.append(f"## Sheet: {sheet.title}")
                rows = self._trim_rows(sheet.iter_rows(values_only=True))
                table_markdown = markdown_table(rows)
                if table_markdown:
                    sections.append(table_markdown)
        finally:
            workbook.close()

        content = compact_blank_lines("\n\n".join(sections))
        return ParseResult(
            content=content,
            metadata={
                "source_type": self.source_type,
                "extraction_tool": self.extraction_tool,
                "extraction_version": self.extraction_version,
                "extraction_params": {"data_only": True},
                "page_count": len(workbook.sheetnames),
                "char_count": len(content),
                "image_count": 0,
            },
        )

    def _trim_rows(self, rows: Any) -> list[list[Any]]:
        materialized = [list(row) for row in rows]
        while materialized and all(value is None for value in materialized[-1]):
            materialized.pop()
        if not materialized:
            return []

        width = max(
            (
                index + 1
                for row in materialized
                for index, value in enumerate(row)
                if value is not None
            ),
            default=0,
        )
        return [row[:width] for row in materialized]
